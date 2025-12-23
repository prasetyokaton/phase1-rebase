import streamlit as st
import pandas as pd
import time
from datetime import datetime
import joblib
import re
import gspread
from google.oauth2.service_account import Credentials
from pathlib import Path
import os
import logging
from pytz import timezone
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from column_setup_config import COLUMN_SETUP_CONFIG
from column_order_config import COLUMN_ORDER_CONFIG
import unicodedata
import math



# Load the model and vectorizer for gender prediction
class GenderPredictor:
    def __init__(self):
        model_path = 'path_files/file1.pkl'
        vectorizer_path = 'path_files/file2.pkl'
        self.model = joblib.load(model_path)
        self.vectorizer = joblib.load(vectorizer_path)
        self.labels = {1: "male", 0: "female"}

    def predict(self, name: str):
        vector = self.vectorizer.transform([name])
        result = self.model.predict(vector)[0]
        proba = self.model.predict_proba(vector).max()
        return self.labels[result], round(proba * 100, 2)


@st.cache_resource(ttl=3600) #Cache berlaku selama 1 jam menit
def load_google_sheets_data():
    #this apply for local
    #if os.path.exists('/app/.secretcontainer/insightsautomation-460807-acdad1ee7590.json'):
    #    SERVICE_ACCOUNT_FILE = '/app/.secretcontainer/insightsautomation-460807-acdad1ee7590.json'
    #else:
    #    SERVICE_ACCOUNT_FILE = './.secretcontainer/insightsautomation-460807-acdad1ee7590.json'

    SERVICE_ACCOUNT_FILE = '/app/.secretcontainer/insightsautomation-460807-acdad1ee7590.json'
    SCOPES = ['https://www.googleapis.com/auth/spreadsheets.readonly']
    creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    client = gspread.authorize(creds)

    spreadsheet_id = "1oL9rAR_SuVOjqOLVnYhr7jp8_qxVXuqc_9Imi3RjEEM"
    spreadsheet = client.open_by_key(spreadsheet_id)

    data = {
        "df_project_list": pd.DataFrame(spreadsheet.worksheet("Project List").get_all_records()),
        #"df_column_setup": pd.DataFrame(spreadsheet.worksheet("Column Setup").get_all_records()),
        "df_rules": pd.DataFrame(spreadsheet.worksheet("Rules").get_all_records()),
        "df_column_order": pd.DataFrame(spreadsheet.worksheet("Column Order Setup").get_all_records()),
        "df_method_1_keyword": pd.DataFrame(spreadsheet.worksheet("Method 1 Keyword").get_all_records()),
        "df_method_selection": pd.DataFrame(spreadsheet.worksheet("Method Selection").get_all_records()),
        "df_official_account_setup": pd.DataFrame(spreadsheet.worksheet("Official Account Setup").get_all_records()),
        "last_updated": spreadsheet.worksheet("NOTES").cell(1, 2).value,
        "df_hashtag_priority" : pd.DataFrame(spreadsheet.worksheet("Hashtag Priority").get_all_records())
    }
    return data


def apply_creator_type_logic(df):
    try:
        # Setup Google Sheets access
        SERVICE_ACCOUNT_FILE = '/app/.secretcontainer/insightsautomation-460807-acdad1ee7590.json'
        SCOPES = ['https://www.googleapis.com/auth/spreadsheets.readonly']
        creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
        client = gspread.authorize(creds)

        # --- Load GSheet for Media Tier & PAP/PA ---
        sheet_main = client.open_by_key("1QZF9yFyI-Bc67yp7hT4pYIAfmrZi1e4VIFsQ7WbIcII")
        df_online = pd.DataFrame(sheet_main.worksheet("Online with AVE - Updated").get_all_records())

        # --- Load GSheet for KOL & HM ---
        sheet_kol = client.open_by_key("1FKIw9tpwiZs2VlIx4xjwPp0u_8BbxFRYF5uY8Jf_ozg")
        df_kol = pd.DataFrame(sheet_kol.worksheet("List KOL Nojorono").get_all_records())
        df_hm = pd.DataFrame(sheet_kol.worksheet("List Homeless Media").get_all_records())

        # Bersihkan kolom
        df_online.columns = df_online.columns.str.strip()
        df_kol.columns = df_kol.columns.str.strip()
        df_hm.columns = df_hm.columns.str.strip()
        
        if "Creator Type" not in df.columns:
            df['Creator Type'] = ""


        # --- STEP 1: PA (Publisher Article) ---
        for idx, row in df[df["Channel"].str.strip().str.lower() == "online media"].iterrows():
            media_name = str(row.get("Media Name", "")).strip()
            media_tier = str(row.get("Media Tier", "")).strip()
            if media_name and media_tier in ["1", "2", "3"]:
                df.at[idx, "Creator Type"] = f"PA Tier {media_tier}"

        # --- STEP 2: PAP (PA Publisher) ---
        for idx, row in df[df["Channel"].str.lower().isin(["instagram", "tiktok"])].iterrows():
            channel = row["Channel"].lower()
            author = str(row.get("Author", "")).strip().lower()
            if channel == "instagram":
                kol = df_online[df_online["Instagram Author Name"].str.strip().str.lower() == author]
            elif channel == "tiktok":
                kol = df_online[df_online["Tiktok Author Name"].str.strip().str.lower() == author]
            else:
                continue
            if not kol.empty:
                tier = str(kol.iloc[0].get("Media Tier", "")).strip()
                if tier in ["1", "2", "3"]:
                    df.at[idx, "Creator Type"] = f"PAP Tier {tier}"

        # --- STEP 3: HM (Homeless Media + Keyword Matching) ---
        media_keywords = [
            "media", "news", "update", "daily", "info", "portal", "tribun", "detik",
            "times", "today", "berita", "channel", "tv", "net", "kompas", 
            "kabar", "indozone", "cnn", "liputan", "official", "forum", "post", "koran",
            "zona", "zone", "idn", "music", "musik", "tech", "health", "headline", "newsroom",
            "film", "local", "lokal", "hospital", "resort", "radio", "digital", 
            "agency", "studio", "group", "restaurant", "university", "universitas", "collage", "grup", "festival",
            "foundation", "institut", "institute", "org", "academy", "school", "sekolah", "kampus", 
            "community", "komunitas", "cafe", "coffee", "sport",
            "wisata", "tempat", "clinic", "fotografi", "videografi", "editor", "visual",
            "jakarta", "bandung", "bali", "jogja", "banten", "surabaya", "nusantara", "bogor", "bekasi", "tangerang",
            "mart", "indomie", "minuman", "oleholeh", "jajanan", "distro", "merch", "grosir", "wholesale", "toserba", "minimarket", "kios",
            "travel", "trip", "explore", "vacation", "holiday", "homestay", "penginapan", "villa", "hotel", "kost", "guesthouse",
            "engineering", "arsitek", "kontraktor", "interior", "furnitur", "elektronik", "mesin",
            "diskon", "promo", "gratis", "reseller", "dropship", "seller", "jualan", "toko",
            "quotes", "motivasi", "edukasi", "trivia"
        ]



        for idx, row in df[df["Channel"].str.lower().isin(["instagram", "tiktok", "facebook", "youtube", "twitter"])].iterrows():
            author = str(row.get("Author", "")).strip().lower()
            match_list = df_hm[df_hm["Author Name"].str.strip().str.lower() == author]

            followers = row.get("Followers", "")
            try:
                followers = int(str(followers).replace(",", "").strip())
            except:
                followers = 0

            is_hm_from_keyword = any(kw in author for kw in media_keywords)
            is_hm_listed = not match_list.empty

            if is_hm_listed or is_hm_from_keyword:
                if followers > 1_000_000:
                    df.at[idx, "Creator Type"] = "HM Mega"
                elif followers > 100_000:
                    df.at[idx, "Creator Type"] = "HM Macro"
                elif followers > 10_000:
                    df.at[idx, "Creator Type"] = "HM Micro"
                elif followers > 1_000:
                    df.at[idx, "Creator Type"] = "HM Nano"
                else:
                    df.at[idx, "Creator Type"] = "HM Micro Nano"

        # --- STEP 4: KOL (Only TikTok, Instagram, Twitter, Facebook, YouTube) ---
        for idx, row in df[
            (df["Channel"].str.lower().isin(["tiktok", "instagram", "youtube", "twitter", "facebook"])) &
            (df["Creator Type"] == "")
        ].iterrows():
            channel = row["Channel"].lower()
            author = str(row.get("Author", "")).strip().lower()
            followers = row.get("Followers", 0)
            try:
                followers = int(str(followers).replace(",", "").strip())
            except:
                followers = 0

            if channel == "tiktok":
                kol_check = df_kol["Author Name Tiktok"].str.strip().str.lower() == author
            elif channel == "instagram":
                kol_check = df_kol["Author Name Instagram"].str.strip().str.lower() == author
            else:
                kol_check = pd.Series([False] * len(df_kol))  # Tidak dicek

            if not kol_check.any() and followers > 0:
                if followers > 1_000_000:
                    df.at[idx, "Creator Type"] = "KOL Mega"
                elif followers > 100_000:
                    df.at[idx, "Creator Type"] = "KOL Macro"
                elif followers > 10_000:
                    df.at[idx, "Creator Type"] = "KOL Micro"
                #elif followers > 1_000:
                else:
                    df.at[idx, "Creator Type"] = "KOL Nano"
                #else:
                #    df.at[idx, "Creator Type"] = "KOL Micro Nano"

    except Exception as e:
        st.error(f"❌ Gagal memproses Creator Type: {e}")
        return df
    
    return df

# Function to process and fill gender prediction if confidence > 80%
def fill_gender(df):
    predictor = GenderPredictor()
    for index, row in df[df['Gender'].isna()].iterrows():
        name = row['Author']  # Assuming there's an 'Author' column
        
        # Skip rows where the 'Author' name is NaN
        if pd.isna(name):
            continue

        # Remove non-alphabetic characters (keep only letters)
        name_cleaned = re.sub(r'[^a-zA-Z]', '', name)

        # Skip empty names after cleaning
        if not name_cleaned:
            continue
        
        # Predict gender
        gender, probability = predictor.predict(name_cleaned)
        
        # Fill the 'Gender' column if the prediction confidence is greater than 70%
        if probability > 80:
            df.at[index, 'Gender'] = gender
    return df


# === Apply Media Tier Logic ===
def apply_media_tier_logic(df):
    try:
        # === SETUP GOOGLE SHEETS API ===
        #di docker dia gabisa /home tp /app/
        SERVICE_ACCOUNT_FILE = '/app/.secretcontainer/insightsautomation-460807-acdad1ee7590.json'
        SCOPES = ['https://www.googleapis.com/auth/spreadsheets.readonly']
        creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
        client = gspread.authorize(creds)

        # === BUKA GOOGLE SHEET ===
        sheet_id = "1QZF9yFyI-Bc67yp7hT4pYIAfmrZi1e4VIFsQ7WbIcII"
        spreadsheet = client.open_by_key(sheet_id)

        # === LOAD kedua sheet ===
        data_client = spreadsheet.worksheet("Le Minerale - from Client").get_all_records()
        data_online = spreadsheet.worksheet("Online with AVE - Updated").get_all_records()
    
        # Buat dictionary media_tier dengan key lowercase
        media_tier_dict_client = {
            row.get("Media Name", "").strip().lower(): row.get("Media Tier")
            for row in data_client if row.get("Media Name")
        }
        media_tier_dict_online = {
            row.get("Media Name", "").strip().lower(): row.get("Media Tier")
            for row in data_online if row.get("Media Name")
        }

        # STEP 1: isi dari client sheet (hanya jika Media Name tidak kosong)
        for index, row in df[df["Media Tier"].isna() | (df["Media Tier"] == "")].iterrows():
            media_name = str(row.get("Media Name", "")).strip().lower()
            if media_name and media_name in media_tier_dict_client:
                df.at[index, "Media Tier"] = media_tier_dict_client[media_name]


        # STEP 2: isi dari online sheet
        for index, row in df[df["Media Tier"].isna() | (df["Media Tier"] == "")].iterrows():
            media_name = str(row["Media Name"]).strip().lower()
            if media_name and media_name in media_tier_dict_online:
                df.at[index, "Media Tier"] = media_tier_dict_online[media_name]


        # STEP 3: isi berdasarkan Ad Value
        for index, row in df[df["Media Tier"].isna() | (df["Media Tier"] == "")].iterrows():
            media_name = str(row["Media Name"]).strip()
            #ad_value = row.get("Ad Value")
            raw_val = row.get("Ad Value")
            ad_value  = to_float(raw_val)

            # Hanya proses jika media_name dan ad_value tidak kosong
            if media_name and media_name.strip() != "" and pd.notna(ad_value):
                try:
                    ad_value = float(ad_value)
                    if ad_value >= 18000000:
                        df.at[index, "Media Tier"] = 1
                    elif ad_value >= 12600000:
                        df.at[index, "Media Tier"] = 2
                    else:
                        df.at[index, "Media Tier"] = 3
                except Exception as e:
                    logging.warning(f"Error parsing ad_value '{ad_value}': {e}")
                    continue  # jika ad_value tidak bisa dikonversi, skip
        return df
    
    except Exception as e:
        st.error(f"❌ Gagal load Google Sheet Media Tier: {e}")
        return df  # keluar tanpa ubah jika gagal
        
        
# Setup logging with dynamic filename
def clean_filename(text):
    return re.sub(r"[^\w\-]", "_", str(text).strip())

def init_logging(project_name, uploaded_filename):
    log_dir = Path("logs"); log_dir.mkdir(exist_ok=True)

    wib = timezone("Asia/Jakarta")
    timestamp = datetime.now(wib).strftime("%Y%m%d_%H%M%S")

    log_filename = f"{timestamp}_{clean_filename(project_name)}_{clean_filename(uploaded_filename)}.log"
    log_path = log_dir / log_filename

    # Gunakan logger khusus modul
    logger = logging.getLogger(__name__)
    logger.setLevel(logging.INFO)

    # Tambah handler **hanya jika belum ada FileHandler** agar tidak dobel saat rerun
    if not any(isinstance(h, logging.FileHandler) for h in logger.handlers):
        fh = logging.FileHandler(log_path, mode="w")
        fmt = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
        fh.setFormatter(fmt)
        logger.addHandler(fh)

    logger.info("🚀 Streamlit app started")
    return log_filename


# === MODELED REACH (MR) — alpha*(Views_used*u) + (1-alpha)*(Followers*r)
# Views_used = Views jika >0; jika 0/kosong → pakai Buzz
# MR harus < Views (bila Views>0) dan jika Views=0 → MR < Buzz


# tetap pakai MR_PARAMS kamu
MR_PARAMS = {
    "instagram": {"alpha": 0.6, "u": 0.7,  "r": 0.035},
    "tiktok":    {"alpha": 0.9, "u": 0.7,  "r": 0.026},
    "youtube":   {"alpha": 0.9, "u": 0.7,  "r": 0.0112},
    "facebook":  {"alpha": 0.5, "u": 0.7,  "r": 0.0015},
    "twitter":   {"alpha": 0.7, "u": 0.6,  "r": 0.00029},
    "x":         {"alpha": 0.7, "u": 0.6,  "r": 0.00029},
    "default":   {"alpha": 0.7, "u": 0.7,  "r": 0.01},
}

def add_mr(df, params=MR_PARAMS):
    df = df.copy()

    # ambil Views numeric (boleh kosong)
    views = (
        pd.to_numeric(df.get("Views", 0).apply(to_float), errors="coerce")
        .fillna(0)
    )
    # jika Views 0/kosong → pakai Buzz
    buzz = (
        pd.to_numeric(df.get("Buzz", 0).apply(to_float), errors="coerce")
        .fillna(0)
    )
    views_eff = views.where(views > 0, buzz)

    # followers
    followers = (
        pd.to_numeric(df.get("Followers", 0).apply(to_float), errors="coerce")
        .fillna(0)
    )

    # hitung MR per baris
    def _mr_row(i, row):
        ch = str(row.get("Channel", "")).strip().lower()
        p  = params.get(ch, params["default"])
        alpha, u, r = p["alpha"], p["u"], p["r"]

        v_eff = views_eff.iat[i]
        foll  = followers.iat[i]

        mr_raw = alpha * (v_eff * u) + (1 - alpha) * (foll * r)

        # round up (ceil), integer
        mr_int = int(math.ceil(mr_raw)) if mr_raw > 0 else 0

        # cap ketat:
        # - kalau Views > 0: MR < Views
        # - kalau Views == 0: MR < Buzz
        cap_base = views.iat[i] if views.iat[i] > 0 else buzz.iat[i]
        if cap_base > 0:
            mr_int = min(mr_int, max(int(cap_base) - 1, 0))
        else:
            mr_int = 0

        return mr_int

    df["MR"] = [ _mr_row(i, row) for i, row in enumerate(df.to_dict("records")) ]
    return df


# Function to update the "Media Tier" visibility in the Column Order Setup sheet
def update_media_tier_visibility(df_column_order):

    # Find the row where 'Column Name' is "Media Tier"
    media_tier_row = df_column_order[df_column_order["Column Name"] == "Media Tier"]

    # Check if "Hide" is "Yes" and change it to "No"
    if not media_tier_row.empty:
        if media_tier_row["Hide"].iloc[0].strip().lower() == "yes":
            df_column_order.loc[df_column_order["Column Name"] == "Media Tier", "Hide"] = "No"
    
    # Return the updated DataFrame
    return df_column_order


# Normalisasi fungsi
def normalize_text(text):
    # Pastikan input adalah string
    text = str(text)
    
    # Menggunakan unicodedata untuk menghapus karakter non-ASCII
    text = unicodedata.normalize('NFD', text)

    # Hapus semua karakter non-ASCII, termasuk simbol dan karakter Unicode khusus
    return ''.join([c for c in text if unicodedata.category(c) != 'Mn'])




# === FUNGSI: Apply Rules ===
def apply_rules(df, rules, output_column, source_output_column):
    import re
    rules.columns = rules.columns.str.strip()
    rules_sorted = rules.sort_values(by="Priority", ascending=False)
    print(f"⚠️ rules_sorted: {rules_sorted}")

    logging.info(f"⚠️ rules_sorted: {rules_sorted}")           

    if output_column not in df.columns:
        df[output_column] = ""

    # Tambahkan: deteksi semua kolom output dari rules
    output_cols_in_rules = [col for col in rules.columns if col.startswith("Output ")]
    for output_col in output_cols_in_rules:
        colname = output_col.replace("Output ", "")
        if colname not in df.columns:
            df[colname] = ""
    priority_tracker = {
        col.replace("Output ", ""): pd.Series([float("inf")] * len(df), index=df.index)
        for col in output_cols_in_rules
    }
    print(f"⚠️ priority_tracker: {priority_tracker}")
    logging.info(f"⚠️ priority_tracker: {priority_tracker}")

    summary_logs = []
    overwrite_tracker = [[] for _ in range(len(df))]
    indeks_awal = 0
    for _, rule in rules_sorted.iterrows():
        col = rule["Matching Column"]
        val = rule["Matching Value"]
        match_type = rule["Matching Type"]
        priority = rule["Priority"]
        channel = str(rule.get("Channel", "")).strip().lower()


        logging.info(f"🧪 Matching rule channel: '{channel}'")
        logging.info(f"🧪 Unique df['Channel']: {df['Channel'].unique()}")

        # Filter Channel
        if channel and "Channel" in df.columns:
            channel_mask = df["Channel"].astype(str).str.strip().str.lower() == channel
        else:
            channel_mask = pd.Series([True] * len(df), index=df.index)
        
        logging.info(f"⚠️ indeks_awal: {indeks_awal}")
        logging.info(f"⚠️ channel_mask: {channel_mask}")
        print(f"⚠️ channel_mask: {channel_mask}")
        indeks_awal = indeks_awal + 1

        # Normalisasi rule value
        normalized_rule_value = normalize_text(str(val))



        # Matching Logic
        if "|" in col:  # Pencarian OR
            parts = [p.strip() for p in col.split("|")]
            mask = pd.Series([False] * len(df))  # Awali dengan mask False
            for part in parts:
                if part in df.columns:
                    try:
                        escaped_val = re.escape(normalized_rule_value)  # Escape special regex characters
                        mask |= df[part].astype(str).str.contains(escaped_val, case=False, na=False)
                    except Exception as e:
                        logging.error(f"Error processing column '{part}' with value '{val}': {e}")
                        continue
        elif "+" in col:  # Pencarian AND
            parts = [p.strip() for p in col.split("+")]
            mask = pd.Series([True] * len(df))  # Awal mask dengan True (karena AND)
            for part in parts:
                if part in df.columns:
                    try:
                        escaped_val = re.escape(normalized_rule_value)
                        mask &= df[part].astype(str).str.contains(escaped_val, case=False, na=False)
                    except Exception as e:
                        logging.error(f"Error processing column '{part}' with value '{val}': {e}")
                        continue
        else:  # Pencarian untuk satu kata kunci
            if col not in df.columns:
                continue
            series = df[col].astype(str)
            mask = series.str.contains(normalized_rule_value, case=False, na=False)

        
        print(f"⚠️ series: {series}")
        logging.info(f"⚠️ series: {series}")

        if match_type == "contains":
            def eval_or_group(group_str):
                """Handle OR and NOT in a group like: 'A|!B|C' """
                parts = group_str.split("|")
                submasks = []
                for part in parts:
                    part = part.strip()
                    if part.startswith("!"):
                        keyword = re.escape(part[1:])
                        submasks.append(~series.str.contains(keyword, case=False, na=False))
                    else:
                        keyword = re.escape(part)
                        submasks.append(series.str.contains(keyword, case=False, na=False))
                return pd.concat(submasks, axis=1).any(axis=1)

            def parse_expression(expr):
                """Parse full expression like '(A|B)+(C|!D)' or just 'A|B' """
                expr = expr.strip()
                if '+' in expr:
                    # Split by top-level + only, respecting parentheses
                    parts = []
                    current = ""
                    depth = 0
                    for c in expr:
                        if c == '(':
                            depth += 1
                        elif c == ')':
                            depth -= 1
                        if c == '+' and depth == 0:
                            parts.append(current.strip())
                            current = ""
                        else:
                            current += c
                    parts.append(current.strip())
                    submasks = []
                    for part in parts:
                        submasks.append(parse_expression(part))
                    return pd.concat(submasks, axis=1).all(axis=1)
                else:
                    # Single group or expression: remove parentheses
                    group = expr
                    if group.startswith("(") and group.endswith(")"):
                        group = group[1:-1]
                    return eval_or_group(group)

            try:
                mask = parse_expression(val)
                logging.info(f"🔍 Evaluated expression: {val}")
                print(f"🔍 Evaluated expression: {val}")
            except Exception as e:
                logging.warning(f"❌ Parsing error: {e}")
                continue

           
        elif match_type == "equals":
            mask = series == val
        elif match_type == "greater_than":
            try:
                val_num = float(val)
                series_num = pd.to_numeric(series, errors="coerce")
                mask = series_num > val_num
            except ValueError:
                continue
        elif match_type == "less_than":
            try:
                val_num = float(val)
                series_num = pd.to_numeric(series, errors="coerce")
                mask = series_num < val_num
            except ValueError:
                continue
        elif match_type == "count_contains":
            try:
                keyword, constraint = val.split(":")
                keyword = re.escape(keyword.strip())
                constraint = constraint.strip()
                counts = series.str.lower().str.count(rf"\b{keyword}\b")
                if "max=" in constraint:
                    max_allowed = int(constraint.replace("max=", "").strip())
                    mask = counts <= max_allowed
                elif "min=" in constraint:
                    min_allowed = int(constraint.replace("min=", "").strip())
                    mask = counts >= min_allowed
                else:
                    continue
            except Exception as e:
                print(f"⚠️ Error parsing count_contains rule: {val} - {e}")
                continue
        else:
            continue
        
        logging.info(f"⚠️ mask: {mask}")
        print(f"⚠️ mask: {mask}")

        update_mask = mask & channel_mask
        logging.info(f"⚠️ update_mask: {update_mask}")
        print(f"⚠️ update_mask: {update_mask}")

        # Apply output to relevant columns
        for output_col in output_cols_in_rules:
            out_val = rule.get(output_col)
            colname = output_col.replace("Output ", "")
            if pd.notna(out_val) and colname in df.columns:
                update_condition = mask & channel_mask & (priority_tracker[colname] > priority)

                if update_condition.any():
                    logging.info(f"✅ Rule applied: {rule.to_dict()}")

                df.loc[update_condition, colname] = out_val
                priority_tracker[colname].loc[update_condition] = priority
                
                for idx in update_condition[update_condition].index:
                    matched_kw = _extract_first_match(series.at[idx], val)
                    overwrite_tracker[idx].append(f"{colname} P{priority}: {out_val} ({matched_kw})")



        tmp_update_mask_sum = 0
        # Log summary
        for output_col in output_cols_in_rules:
            out_val = rule.get(output_col)
            colname = output_col.replace("Output ", "")
            if pd.notna(out_val) and colname in df.columns:
                affected_count = mask.sum()
                if affected_count > 0 and pd.notna(out_val):
                    summary_logs.append({
                        "Priority": priority,
                        "Matching Column": col,
                        "Matching Value": val,
                        "Matching Type": match_type,
                        "Channel": channel,
                        "Affected Rows": affected_count,
                        "Output Column": colname,
                        "Output Value": out_val
                    })


    # ---- NEW : build Rules Affected & Rules Affected Words ----
    if output_column == "Noise Tag":
        rules_affected_counts = []
        rules_affected_words  = []

        for overwrites in overwrite_tracker:
            # jumlah rule (tanpa pembagi 3)
            rules_affected_counts.append(len(overwrites))

            words_fmt = []
            for item in overwrites:
                # pola: "<Column> P<prio>: <out_val> (<keyword>)"
                m = re.match(r'(.+?) P(\d+): (.+?) \((.+?)\)', item)
                if m:
                    _, _, out_val, keyword = m.groups()
                    words_fmt.append(f"{out_val} ({keyword})")
            rules_affected_words.append("|".join(words_fmt))

        df["Rules Affected"]       = rules_affected_counts
        df["Rules Affected Words"] = rules_affected_words


    summary_df = pd.DataFrame(summary_logs)
    return df, summary_df


#Untuk menentukan official account
def apply_official_account_logic(df, setup_df, project_name):
    import re
    setup_df.columns = setup_df.columns.str.strip()
    
    # Ubah nilai TRUE/FALSE jadi Yes/No (string)
    setup_df["Verified Account"] = setup_df["Verified Account"].apply(
        lambda x: "yes" if str(x).strip().lower() in ["true", "yes", "1"] else "no"
    )

    # Ambil rules yang sesuai project
    setup_project = setup_df[setup_df["Project"] == project_name]

    for _, row in setup_project.iterrows():
        verified = str(row.get("Verified Account", "")).strip().lower()
        channel = str(row.get("Channel", "")).strip().lower()
        col = row["Matching Column"]
        val = row["Matching Value"]
        match_type = row["Matching Type"]

        if col not in df.columns or "Channel" not in df.columns or "Verified Account" not in df.columns:
            continue

        # Filter: channel dan verified
        mask = (
            df["Verified Account"].astype(str).str.strip().str.lower() == verified
        ) & (
            df["Channel"].astype(str).str.strip().str.lower() == channel
        )

        series = df[col].astype(str)

        if match_type == "contains":
            pattern = re.escape(val)
            mask &= series.str.contains(pattern, case=False, na=False)

        elif match_type == "equals":
            mask &= series == val

        else:
            continue

        df.loc[mask, "Official Account"] = "Official Account"
        df.loc[mask, "Noise Tag"] = "1"
        df.loc[mask, "Sentiment"] = "positive" #ganti semua sentiment OA jadi positive

    return df


# --- Helper: pilih URL sumber sesuai channel -------------
def _pick_source_url(row):
    ch = (row.get("Channel") or "").strip().lower()
    if ch == "youtube":
        return row.get("Channel URL", "")      # prioritas channel URL
    else:
        return row.get("Link URL", "")         # default: Link URL


def _extract_username(link: str, channel: str) -> str:
    link = (link or "").lower()

    patterns = {
        # 1) anything between /@ and the next “/”
        "tiktok"  : r"tiktok\.com/@([^/?#]+)",
        # 2) anything after twitter.com/ until next “/”
        "twitter" : r"twitter\.com/([^/?#]+)",
        # 3) handle /channel/xxxx , /@handle , /user/xxxx
        "youtube" : r"youtube\.com/(?:channel/|@|user/)([^/?#]+)",
    }
    pat = patterns.get(channel.lower())
    if not pat:
        return ""

    m = re.search(pat, link)
    return m.group(1) if m else ""


def apply_official_account_logic_v2(df, setup_df, project_name):
    """
    • TikTok / Twitter / YouTube  ➜ match by username extracted from URL
    • Instagram / Facebook       ➜ match by Author + Verified Account filter
    Returns dataframe with Official Account + Noise Tag updated.
    """
    setup_df = setup_df.copy()
    setup_df.columns = setup_df.columns.str.strip()

    # ↓ keep only rules for this project
    setup_project = setup_df[setup_df["Project"] == project_name]

    # --- Normalise verified flag in sheet to 'yes'/'no'
    setup_project["Verified Account"] = setup_project["Verified Account"].apply(
        lambda x: "yes" if str(x).strip().lower() in ["true", "yes", "1"] else "no"
    )

    # Pre-compute username column once  ← pakai helper baru
    if "Username" not in df.columns:
        df["Username"] = df.apply(
            lambda r: _extract_username(
                _pick_source_url(r),               # ← ganti 1 baris ini
                r.get("Channel", "")
            ),
            axis=1
        ).str.strip().str.lower()


    # Iterate over rules
    for _, rule in setup_project.iterrows():
        channel   = str(rule.get("Channel", "")).strip().lower()
        verified  = str(rule.get("Verified Account", "")).strip().lower()   # 'yes' / 'no'
        target    = str(rule.get("Matching Value", "")).strip().lower()     # official username / author
        if not channel or not target:
            continue

        # --- Build filter mask per channel ----------
        chan_mask = df["Channel"].astype(str).str.strip().str.lower() == channel

        if channel in ["instagram", "facebook"]:
            # use Author + verified check
            col_match = df["Author"].astype(str).str.strip().str.lower() == target
            ver_mask  = df["Verified Account"].astype(str).str.strip().str.lower() == verified
            mask = chan_mask & ver_mask & col_match
        elif channel in ["tiktok", "twitter", "youtube"]:
            col_match = df["Username"] == target
            mask = chan_mask & col_match
        else:
            continue  # unsupported channel

        # --- Apply flags when matched ---------------
        df.loc[mask, "Official Account"] = "Official Account"
        df.loc[mask, "Noise Tag"]        = "1"
        df.loc[mask, "Sentiment"] = "positive" #ganti semua sentiment OA jadi positive

    return df



#def bar progress
def update_progress(step, total_steps, description):
    percent = int((step / total_steps) * 100)
    progress.progress(percent)
    status_text.text(f"{description} ({percent}%)")


# Function to process and fill Hashtag Priority
def assign_hashtag_priority(df, df_hp, project_name):
    # Ambil rule untuk project
    hp_df = (
        df_hp
        .assign(Hashtag=lambda d: d["Hashtag"].str.strip())   # buang spasi
        .query("Project == @project_name")
    )
    if hp_df.empty:
        return df

    for idx, row in df.iterrows():
        hashtags = str(row.get("Hashtag", "")).lower().split("|")
        best = (
            hp_df[hp_df["Hashtag"].str.lower().isin(hashtags)]
            .sort_values("Priority")          # kecil = prioritas tinggi
            .head(1)
        )
        df.at[idx, "Hashtag Priority"] = (
            best["Hashtag"].iloc[0] if not best.empty else ""
        )
    return df


#jagain untuk float tidak ada koma atau titik
def to_float(val):
    """
    Konversi string angka Indonesia/Inggris + akhiran k/m:
    - '1.234' -> 1234
    - '1,234' -> 1234
    - '1.234,56' -> 1234.56
    - '1,234.56' -> 1234.56
    - '1,7' -> 1.7
    - '1.7k' / '1,7k' -> 1700
    - '18.000.000' -> 18000000
    - '18,000,000' -> 18000000
    """
    if val is None:
        return None
    s = str(val).strip().lower()
    if s in ("", "nan", "none", "-"):
        return None

    # handle suffix
    mult = 1
    if s.endswith("k"):
        mult, s = 1_000, s[:-1].strip()
    elif s.endswith("m"):
        mult, s = 1_000_000, s[:-1].strip()

    # keep only digits, comma, dot, minus
    s = re.sub(r"[^0-9,.\-]", "", s)

    # case: both dot and comma exist → tentukan mana desimal
    if "," in s and "." in s:
        # asumsikan format Indonesia: '.' = ribuan, ',' = desimal (contoh: 1.234,56)
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            # format Inggris: ',' = ribuan, '.' = desimal (contoh: 1,234.56)
            s = s.replace(",", "")
    else:
        # hanya koma → anggap koma sebagai desimal
        if "," in s:
            s = s.replace(",", ".")
        # hanya titik → biarkan (sudah desimal Inggris)

    try:
        return float(s) * mult
    except:
        return None



def _extract_first_match(text: str, expr: str) -> str:
    """
    text : isi kolom yang sedang dicek (Content, Title, dst.)
    expr : nilai Matching Value (bisa 'A|B|C' atau ekspresi '(A|B)+(C|D)')
    Return  : satu keyword yang pertama kali ketemu di text.
              Jika tidak ketemu apa-pun, fallback ke expr apa adanya.
    """
    # buang kurung luar
    expr = expr.strip()
    if expr.startswith("(") and expr.endswith(")"):
        expr = expr[1:-1]

    # pisah berdasarkan "|" di level teratas (abaikan + dan sub-parentheses)
    depth = 0
    token = ""
    tokens = []
    for c in expr:
        if c == '(':
            depth += 1
        elif c == ')':
            depth -= 1

        if c == '|' and depth == 0:
            tokens.append(token.strip())
            token = ""
        else:
            token += c
    tokens.append(token.strip())

    # bersihkan awalan "!"  → hanya keyword positif
    tokens = [t.lstrip('!').strip() for t in tokens if t]

    # cari token yang benar-benar ada di text
    text_low = text.lower()
    for t in tokens:
        if t and t.lower() in text_low:
            return t
    return tokens[0] if tokens else expr        # fallback










# === MULAI STREAMLIT APP ===
st.title("Insight Automation Phase 1")

#Load json
google_sheet_data = load_google_sheets_data()

if google_sheet_data is None:
    st.stop()

df_project_list = google_sheet_data["df_project_list"]
#df_column_setup = google_sheet_data["df_column_setup"] => tidak terpakai karena tidak lagi ambil dari google sheet
df_rules = google_sheet_data["df_rules"]
#df_column_order = google_sheet_data["df_column_order"]
df_method_1_keyword = google_sheet_data["df_method_1_keyword"]
df_method_selection = google_sheet_data["df_method_selection"]
df_official_account_setup = google_sheet_data["df_official_account_setup"]
last_updated = google_sheet_data["last_updated"]

#====== SUDAH DENGAN AMBIL DARI FILE =========
# NEW – pull from local config
column_setup_df = pd.DataFrame(COLUMN_SETUP_CONFIG)
df_column_order = pd.DataFrame(COLUMN_ORDER_CONFIG)




st.markdown("#### Pilih Project Name:")
#st.caption(f"📄 Rules terakhir diperbarui pada: {last_updated}")
col1, col2 = st.columns([6, 1])  # Kolom kecil untuk tombol

with col1:
    st.caption(f"📄 Rules terakhir diperbarui pada: {last_updated or '(tidak tersedia)'}")

with col2:
    if st.button("🔄", help="Refresh data dari Google Sheets (paksa ambil ulang)"):
        st.cache_resource.clear()  # Hapus semua cache resource (termasuk load_google_sheets_data)
        st.rerun()      # Refresh halaman


project_name = st.selectbox("", ["Pilih Project"] + df_project_list["Project Name"].dropna().tolist())

# === DEBUG AREA UNTUK TESTING RULES MANUAL ===
DEBUG_MODE = False  # Ubah ke False atau comment saat production

if DEBUG_MODE and project_name != "Pilih Project":
    st.markdown("---")
    st.markdown("### 🧪 Debug Rule Testing (Developer Only)")
    test_content = st.text_area("✍️ Masukkan contoh isi konten untuk testing:")

    if st.button("🔍 Tes Rules untuk Konten Ini"):
        # Ambil rules sesuai project
        rules_default = df_rules[df_rules["Project"] == "Default"]
        rules_project = df_rules[df_rules["Project"] == project_name] if project_name in df_rules["Project"].values else pd.DataFrame()
        rules_combined = pd.concat([rules_default, rules_project], ignore_index=True)

        # Simulasi single row DataFrame
        df_test = pd.DataFrame([{
            "Channel": "instagram",  # bisa kamu buat dropdown juga
            "Content": test_content,
            "Title": "",
            "Campaigns": project_name,
            "Verified Account": "No",
            "Media Name": "",
            "Followers": 0,
            "Author": "testauthor"
        }])

        # Apply rules
        df_test, _ = apply_rules(df_test, rules_combined, output_column="Noise Tag", source_output_column="Output Noise Tag")
        df_test, _ = apply_rules(df_test, rules_combined, output_column="Issue", source_output_column="Output Issue")
        df_test, _ = apply_rules(df_test, rules_combined, output_column="Sub Issue", source_output_column="Output Sub Issue")

        st.success("✅ Hasil Evaluasi Rules:")
        st.write(f"**Noise Tag**: {df_test.at[0, 'Noise Tag']}")
        st.write(f"**Issue**: {df_test.at[0, 'Issue']}")
        st.write(f"**Sub Issue**: {df_test.at[0, 'Sub Issue']}")


uploaded_raw = st.file_uploader("Upload Raw Data", type=["xlsx"], key="raw")

remove_duplicate_links = st.checkbox("Remove duplicate link")
keep_raw_data = st.checkbox("Keep RAW Data (Save original file as separate sheet)")

# New checkboxes for Media Tier
apply_media_tier = st.checkbox("Apply Media Tier")

if not apply_media_tier:
    st.caption("🔒 Centang 'Apply Media Tier' terlebih dahulu untuk mengaktifkan Creator Type")
    st.checkbox("Apply Creator Type", key="disabled_creator_type", disabled=True)
    apply_creator_type = False
else:
    apply_creator_type = st.checkbox("Apply Creator Type")
    if apply_creator_type:
        if (
            df_column_order is not None and
            "Column Name" in df_column_order.columns and
            "Hide" in df_column_order.columns
        ):
            df_column_order.loc[
                df_column_order["Column Name"] == "Creator Type", "Hide"
            ] = "No"


submit = st.button("Submit")

if submit:
    # Progress bar setup
    progress = st.progress(0)
    status_text = st.empty()

    if project_name == "Pilih Project" or uploaded_raw is None:
        st.error("❌ Anda harus memilih project dan upload raw data sebelum submit.")
    else:
        #st.success(f"✅ Project: {project_name} | File Loaded Successfully!")
        
        # Init logging
        log_filename = init_logging(project_name, uploaded_raw.name)
        logging.info(f"📂 Project: {project_name}")
        logging.info(f"📎 Uploaded File: {uploaded_raw.name}")

        #Load data berhasil
        update_progress(1, 14, "📁 Load File ..")
        time.sleep(1)
        start_time = time.time()

        df_raw = pd.read_excel(uploaded_raw, sheet_name=0)
        if "Campaign" in df_raw.columns:
            df_raw = df_raw.rename(columns={"Campaign": "Campaigns"})
        df_processed = df_raw.copy()

        # Remove duplicate link
        if remove_duplicate_links and "Link URL" in df_processed.columns:
            before_count = len(df_processed)
            df_processed = df_processed.drop_duplicates(subset="Link URL").reset_index(drop=True)
            after_count = len(df_processed)
            st.info(f"🔁 Removed {before_count - after_count} duplicate rows based on 'Link URL'")

        #Standardize Verified Account
        update_progress(2, 14, "📁 Standardize Verified Account ..")
        time.sleep(1)
        # Standardize Verified Account
        if "Verified Account" in df_processed.columns:
            df_processed["Verified Account"] = (
                df_processed["Verified Account"].astype(str).str.strip().str.lower().replace({"-": "no", "": "no", "nan": "no"})
            )
            df_processed["Verified Account"] = df_processed["Verified Account"].apply(lambda x: "Yes" if x == "yes" else "No")



        #Create kolom hashtag
        update_progress(3, 14, "📁 Finding Hashtag ..")
        time.sleep(1)
        # === Tambah kolom Hashtag ===
        if "Hashtag" not in df_processed.columns:
            # Sisipkan setelah kolom Content jika ada
            insert_at = (
                df_processed.columns.get_loc("Content") + 1
                if "Content" in df_processed.columns
                else len(df_processed.columns)
            )
            df_processed.insert(loc=insert_at, column="Hashtag", value="")

        df_processed["Hashtag"] = (
            df_processed["Content"]
            .astype(str)
            .str.findall(r"#\w+")
            .apply(lambda tags: "|".join(tags))
        )


        # Memanggil fungsi untuk menetapkan Hashtag Priority
        if "Hashtag" in df_processed.columns:
            #df_processed = assign_hashtag_priority(df_processed, df_rules, project_name)
            
            df_processed = assign_hashtag_priority(
                df_processed,
                google_sheet_data["df_hashtag_priority"],
                project_name
            )

        # Pastikan kolom "Hashtag Priority" berada setelah kolom "Hashtag"
        if "Hashtag Priority" not in df_processed.columns:
            df_processed["Hashtag Priority"] = ""

        # Menyusun kolom agar "Hashtag" dan "Hashtag Priority" berada berdampingan
        columns = list(df_processed.columns)
        if "Hashtag" in columns and "Hashtag Priority" in columns:
            hashtag_index = columns.index("Hashtag")
            hashtag_priority_index = columns.index("Hashtag Priority")
            
            # Pastikan Hashtag Priority berada setelah Hashtag
            if hashtag_index < hashtag_priority_index:
                columns.insert(hashtag_index + 1, columns.pop(hashtag_priority_index))
            else:
                columns.insert(hashtag_index + 1, "Hashtag Priority")

        df_processed = df_processed[columns]



        # Setup Columns
        update_progress(4, 14, "📁 Setup Columns ..")
        time.sleep(1)
        
        #==== sudah tidak terpakai karena tidak ambil dari google sheet lagi"
        # df_column_setup sudah diambil dari google_sheet_data (lihat langkah 1)
        #column_setup_default = df_column_setup[df_column_setup["Project"] == "Default"]
        #column_setup_project = df_column_setup[df_column_setup["Project"] == project_name]

        #column_setup_combined = pd.concat([column_setup_default, column_setup_project], ignore_index=True)

        

        # Default + (optional) project-specific rows
        column_setup_combined = pd.concat([
            column_setup_df[column_setup_df["Project"] == "Default"],
            column_setup_df[column_setup_df["Project"] == project_name]
        ]).reset_index(drop=True)




        for _, row in column_setup_combined.iterrows():
            col = row["Target Column"]
            ref_col = row["Reference Column"]
            pos = row["Position"]
            default = row["Default Value"] if pd.notna(row["Default Value"]) else ""

            if col not in df_processed.columns:
                # Jika kolom belum ada, tambahkan dan isi default
                if ref_col in df_processed.columns:
                    ref_idx = df_processed.columns.get_loc(ref_col)
                    insert_at = ref_idx if pos == "before" else ref_idx + 1
                    df_processed.insert(loc=insert_at, column=col, value=default)
                else:
                    df_processed[col] = default
            else:
                # Jika kolom sudah ada, isi semua nilai kosong / NaN dengan default
                df_processed[col] = df_processed[col].fillna("").replace("", default)

        


        # Bersihkan trailing .0 hanya untuk kolom 'Noise Tag' jika diperlukan
        if "Noise Tag" in df_processed.columns and df_processed["Noise Tag"].notna().any():
            try:
                series_str = df_processed["Noise Tag"].astype(str)
                if any(series_str.dropna().str.contains(r"\.0$", regex=True, na=False)):
                    df_processed["Noise Tag"] = series_str.replace({r"\.0$": ""}, regex=True)
            except Exception as e:
                st.warning(f"⚠️ Gagal membersihkan kolom 'Noise Tag': {e}")


        # Load Rules
        update_progress(5, 14, "📁 Load Rules ..")
        time.sleep(1)
        # === Apply Rules ===
        rules_default = df_rules[df_rules["Project"] == "Default"]
        rules_project = df_rules[df_rules["Project"] == project_name] if project_name in df_rules["Project"].values else pd.DataFrame()

        rules_combined = pd.concat([rules_default, rules_project], ignore_index=True)
        print(f"⚠️ rules_combined: {rules_combined}")


        # Apply untuk Noise Tag
        update_progress(6, 14, "⚙️ Apply Rules Noise Tag ..")
        df_processed, summary_df = apply_rules(
            df=df_processed,
            rules=rules_combined,
            output_column="Noise Tag",
            source_output_column="Output Noise Tag"
        )

        # Apply Gender Prediction
        #df_processed = fill_gender(df_processed)

        # Tambahkan ini untuk Issue
        update_progress(7, 14, "⚙️ Apply Rules Issue ..")
        df_processed, summary_df_issue = apply_rules(
            df=df_processed,
            rules=rules_combined,
            output_column="Issue",
            source_output_column="Output Issue"
        )

        update_progress(8, 14, "⚙️ Apply Rules Sub Issue ..")
        df_processed, summary_df_sub_issue = apply_rules(
            df=df_processed,
            rules=rules_combined,
            output_column="Sub Issue",
            source_output_column="Output Sub Issue"
        )


        # Gabungkan summary Noise Tag + Issue + Sub Issue
        summary_combined = pd.concat([summary_df, summary_df_issue, summary_df_sub_issue], ignore_index=True)

        # hitung followers
        update_progress(9, 14, "⚙️ Count Followers and Modelled Reach ..")
        time.sleep(1)
        # === Hitung kolom Followers ===
        if "Original Reach" in df_processed.columns and "Potential Reach" in df_processed.columns:
            df_processed["Followers"] = df_processed["Original Reach"].fillna(0) + df_processed["Potential Reach"].fillna(0)

        # === Hitung MR ===
        df_processed.drop(columns=["MR"], errors="ignore", inplace=True)
        df_processed = add_mr(df_processed)



        # Apply Official Account Logic dari setup sheet
        #df_processed = apply_official_account_logic(df_processed, df_official_account_setup, project_name)
        # Baru ― v2
        # Apply Official Accoun
        update_progress(10, 14, "⚙️ Apply Official Account ..")
        time.sleep(1)
        df_processed = apply_official_account_logic_v2(df_processed, df_official_account_setup, project_name)

        #st.write("🔍 DEBUG - cek nilai df_processed.columns:", list(df_processed.columns) if df_processed is not None else "df_processed is None")
        

        # Load Rules
        update_progress(11, 14, "⚙️ Apply Official Account ..")
        time.sleep(1)
        # Apply Media Tier if checked
        if apply_media_tier:
            # Apply Media Tier logic
            df_processed = apply_media_tier_logic(df_processed)

            # Update "Media Tier" visibility to be shown
            df_column_order = update_media_tier_visibility(df_column_order)

        #st.write("🔍 DEBUG after media tier - cek nilai df_processed.columns:", list(df_processed.columns) if df_processed is not None else "df_processed is None")
        

        # Load Rules
        update_progress(12, 14, "⚙️ Apply Creator Type, Media Tier, Column Order and Affected Rules ..")
        time.sleep(1)

        # Pastikan Creator Type juga ditampilkan (tidak di-hide)
        if "Creator Type" in df_column_order["Column Name"].values:
            creator_type_row = df_column_order[df_column_order["Column Name"] == "Creator Type"]
            if not creator_type_row.empty and creator_type_row["Hide"].iloc[0].strip().lower() == "yes":
                df_column_order.loc[df_column_order["Column Name"] == "Creator Type", "Hide"] = "No"

        #kalau creator type nyala
        if apply_creator_type:
            df_processed = apply_creator_type_logic(df_processed)

        #st.write("🔍 DEBUG after creator type - cek nilai df_processed.columns:", list(df_processed.columns) if df_processed is not None else "df_processed is None")
        
        # Setup Column Order
        if project_name in df_column_order["Project"].values:
            ordered_cols = df_column_order[df_column_order["Project"] == project_name]
        else:
            ordered_cols = df_column_order[df_column_order["Project"] == "Default"]


        ordered_cols = ordered_cols[ordered_cols["Hide"].str.lower() != "yes"]["Column Name"].tolist()
        final_cols = [col for col in ordered_cols if col in df_processed.columns]



        # Pastikan Rules Affected dan Rules Affected Words ikut disimpan
        if "Rules Affected" in df_processed.columns and "Rules Affected Words" in df_processed.columns:
            if "Rules Affected" not in final_cols:
                final_cols.append("Rules Affected")
            if "Rules Affected Words" not in final_cols:
                final_cols.append("Rules Affected Words")
        
        if "Creator Type" not in final_cols and "Creator Type" in df_processed.columns:
            final_cols.append("Creator Type")

        # --- sisipkan kolom Hashtag di output tepat setelah "Sub Issue"
        if "Hashtag" in df_processed.columns and "Hashtag" not in final_cols:
            if "Sub Issue" in final_cols:
                insert_at = final_cols.index("Sub Issue") + 1
                final_cols.insert(insert_at, "Hashtag")
            else:
                final_cols.append("Hashtag")

        #df_final = df_processed[final_cols] sekali saja setelah hashtag priorty




        # Pastikan kolom "Hashtag Priority" berada setelah kolom "Hashtag"
        if "Hashtag" in df_processed.columns and "Hashtag Priority" in df_processed.columns:
            hashtag_index = df_processed.columns.get_loc("Hashtag")
            hashtag_priority_index = df_processed.columns.get_loc("Hashtag Priority")
            
            # Pastikan Hashtag Priority berada setelah Hashtag
            if hashtag_index < hashtag_priority_index:
                columns.insert(hashtag_index + 1, columns.pop(hashtag_priority_index))
            else:
                columns.insert(hashtag_index + 1, "Hashtag Priority")

        # Update final_cols with "Hashtag Priority" if necessary
        if "Hashtag Priority" not in final_cols and "Hashtag Priority" in df_processed.columns:
            final_cols.append("Hashtag Priority")

        # --- Sisipkan Hashtag Priority setelah Hashtag di final_cols
        if "Hashtag" in final_cols and "Hashtag Priority" in final_cols:
            final_cols.remove("Hashtag Priority")
            hashtag_index = final_cols.index("Hashtag")
            final_cols.insert(hashtag_index + 1, "Hashtag Priority")



        # Pastikan MR ikut output tepat setelah Followers
        if "MR" in df_processed.columns and "MR" not in final_cols:
            if "Followers" in final_cols:
                final_cols.insert(final_cols.index("Followers") + 1, "MR")
            else:
                final_cols.append("MR")





        # Update final dataframe with the ordered columns
        df_final = df_processed[final_cols]





        # simpan ke output_buffer
        update_progress(13, 14, "📊 Prepare file ..")

        # Save Output
        tanggal_hari_ini = datetime.now().strftime("%Y-%m-%d")
        #output_filename = f"{project_name}_{tanggal_hari_ini}.xlsx"
        output_filename = f"{Path(uploaded_raw.name).stem}_phase1.xlsx"

        # === BUILD Data Collection Overview =========================================
        # === Buat Summary Data (Channel x Official Account) ===
        df_summary_base = df_final.copy()

        # Grouping by Channel and Official Account
        summary_group = df_summary_base.groupby(["Channel", "Official Account"]).size().unstack(fill_value=0)

        # Total official & non-official
        total_official = summary_group.get("Official Account", pd.Series(dtype=int)).sum()
        total_nonofficial = summary_group.get("Non Official Account", pd.Series(dtype=int)).sum()

        # Buat struktur tabel summary
        summary_table = []
        for channel in summary_group.index:
            official_count = summary_group.at[channel, "Official Account"] if "Official Account" in summary_group.columns else 0
            nonofficial_count = summary_group.at[channel, "Non Official Account"] if "Non Official Account" in summary_group.columns else 0

            official_pct = f"{official_count} ({official_count / total_official * 100:.2f}%)" if total_official > 0 else "0 (0.00%)"
            nonofficial_pct = f"{nonofficial_count} ({nonofficial_count / total_nonofficial * 100:.2f}%)" if total_nonofficial > 0 else "0 (0.00%)"

            summary_table.append({
                "Data Source (Channel)": channel,
                "Official Account": official_pct,
                "Non-Official Account": nonofficial_pct
            })

        df_summary = pd.DataFrame(summary_table)
        # --- Tambahkan baris Total di akhir tabel summary ---
        total_row = {
            "Data Source (Channel)": "Total",
            "Official Account": total_official,          # hanya angka total
            "Non-Official Account": total_nonofficial    # hanya angka total
        }
        df_summary = pd.concat([df_summary, pd.DataFrame([total_row])], ignore_index=True)



        # === BUILD SHARE-OF-VOICE TABLE =========================================
        campaigns = sorted(df_final["Campaigns"].dropna().unique())

        # Flag OA / Non-OA
        df_sov_base = df_final.copy()
        df_sov_base["OA Flag"] = df_sov_base["Official Account"].apply(
            lambda x: "Official Account"
            if str(x).strip().lower() == "official account" else "Non Official Account"
        )

        # Pivot hitung jumlah
        pivot = (
            df_sov_base
            .groupby(["Channel", "OA Flag", "Campaigns"])
            .size()
            .unstack(fill_value=0)          # kolom = Campaigns
        )

        # Total per campaign & OA
        tot_off  = df_sov_base[df_sov_base["OA Flag"]=="Official Account"].groupby("Campaigns").size()
        tot_non  = df_sov_base[df_sov_base["OA Flag"]=="Non Official Account"].groupby("Campaigns").size()

        # Susun baris SOV
        channel_order = [
            "Twitter","Facebook","Instagram","TikTok","YouTube",
            "Online Media","Printmedia","Forum","Blog","TV"
        ]
        rows = []
        for ch in channel_order:
            for flag in ["Official Account","Non Official Account"]:
                if (ch, flag) not in pivot.index:
                    counts = {c:0 for c in campaigns}
                else:
                    counts = pivot.loc[(ch,flag)].to_dict()

                row = {
                    "Channel": ch if flag=="Official Account" else "",    # baris kedua channel dikosongkan
                    "Account Type": flag
                }
                for camp in campaigns:
                    count = counts.get(camp, 0)
                    denom = tot_off.get(camp,0) if flag=="Official Account" else tot_non.get(camp,0)
                    pct   = f"{count/denom*100:.2f}%" if denom else "0.00%"
                    row[camp] = f"{count} ({pct})"
                rows.append(row)

        df_sov = pd.DataFrame(rows)

        # Pakai header kosong untuk G2-H2
        df_sov.columns = ["", ""] + campaigns





        ### --- TOP 10 AUTHOR BY BUZZ & BY POST ------------------------------
        # GANTI seluruh blok pembuatan `top_post` di bawah ini

        # 1) by Buzz ────────── (tetap sama)
        top_buzz = (
            df_final
            .groupby("Author", as_index=False)
            .agg({"Buzz": "sum",
                    "Channel": "first",
                    "Content": "first",
                    "Link URL": "first",
                    "Sentiment": "first"})   
            .sort_values("Buzz", ascending=False)
            .head(10)
            .reset_index(drop=True)
        )
        top_buzz.insert(0, "Rank", top_buzz.index + 1)

        # 2) by Post ──────────  ✅ perbaikan kolom
        top_post = (
            df_final
            .groupby("Author", as_index=False)
            .size()                         # hitung baris = total post
            .rename(columns={"size": "Total Post"})
            .sort_values("Total Post", ascending=False)
            .head(10)
            .reset_index(drop=True)
        )
        top_post.insert(0, "Rank", top_post.index + 1)



        # ---------------------------------------------------------------
        # === TOP 10 HASHTAG by Buzz & by Post ==========================
        # ---------------------------------------------------------------
        def _uniq_tags(tag_str):
            """ubah '#a|#b|#a' → {'#a', '#b'}"""
            tags = [t.strip() for t in str(tag_str).split("|") if t.strip()]
            return set(tags)

        # ── explode hashtag ke baris –
        tags_df = (
            df_final[["Hashtag", "Buzz"]]
            .assign(Hashtag=lambda d: d["Hashtag"].apply(_uniq_tags))
            .explode("Hashtag")
            .dropna(subset=["Hashtag"])
        )

        # by Buzz
        top_hash_buzz = (
            tags_df
            .groupby("Hashtag", as_index=False)["Buzz"].sum()
            .rename(columns={"Buzz": "Total Buzz"})
            .sort_values("Total Buzz", ascending=False)
            .head(10)
            .reset_index(drop=True)
        )
        top_hash_buzz.insert(0, "Rank", top_hash_buzz.index + 1)

        # by Post
        top_hash_post = (
            tags_df[["Hashtag"]]
            .value_counts()
            .head(10)
            .reset_index(name="Total Post")
        )
        top_hash_post.insert(0, "Rank", top_hash_post.index + 1)



        # ---------------------------------------------------------------
        # === SENTIMENT PERCENTAGE  &  TOP-POST POS / NEG  ===============   ### NEW
        # ---------------------------------------------------------------
        # Normalisasi nilai Sentiment → lower() untuk jaga-jaga
        df_final["Sentiment"] = (
            df_final["Sentiment"]
            .astype(str)
            .str.strip()
            .str.lower()
        )

        sent_counts = df_final["Sentiment"].value_counts()
        total_sent  = sent_counts.sum()

        sent_pct_rows = [
            {
                "Sentiment": label.capitalize(),
                "Value":     f"{cnt} ({cnt/total_sent*100:.1f}%)"
            }
            for label, cnt in sent_counts.items()
        ]
        df_sent_pct = pd.DataFrame(sent_pct_rows)                       ### NEW

        # ── TOP 10 POSITIVE / NEGATIVE by Buzz ─────────────────────────
        #top_posbuzz = (
        #    df_final[df_final["Sentiment"]=="positive"]
        #    .sort_values("Buzz", ascending=False)
        #    .head(10)
        #    .loc[:, ["Author","Channel","Content","Link URL","Buzz"]]
        #    .reset_index(drop=True)
        #)
        #top_posbuzz.insert(0, "No", top_posbuzz.index + 1)

        

        # === TOP 10 POSITIVE POST  (OA vs Non-OA) =====================
        is_pos   = df_final["Sentiment"].str.lower().eq("positive")
        is_oa    = df_final["Official Account"].str.strip().str.lower() == "official account"
        is_nonoa = ~is_oa                                        # semua selain OA

        def _top(df):
            t = (
                df.sort_values("Buzz", ascending=False)
                .head(10)
                .loc[:, ["Author", "Channel", "Content", "Link URL", "Buzz"]]
                .reset_index(drop=True)
            )
            t.insert(0, "No", t.index + 1)
            return t

        top_posbuzz_oa   = _top(df_final[is_pos & is_oa])
        top_posbuzz_nona = _top(df_final[is_pos & is_nonoa])


        # === TOP 10 NEGATIVE POST ======
        top_negbuzz = (
            df_final[df_final["Sentiment"]=="negative"]
            .sort_values("Buzz", ascending=False)
            .head(10)
            .loc[:, ["Author","Channel","Content","Link URL","Buzz"]]
            .reset_index(drop=True)
        )
        top_negbuzz.insert(0, "No", top_negbuzz.index + 1)




        # ---------------------------------------------------------------
        # === TOP CONTENT BY BUZZ =======================================
        # ---------------------------------------------------------------
        top_content_buzz = (
            df_final
            .sort_values("Buzz", ascending=False)
            .head(10)
            .loc[:, ["Author", "Channel", "Content", "Link URL", "Buzz", "Sentiment"]]
            .reset_index(drop=True)
        )
        top_content_buzz.insert(0, "No", top_content_buzz.index + 1)

        # ---------------------------------------------------------------
        # === TOP CAMPAIGNS PERFORMANCE ================================
        # ---------------------------------------------------------------
        perf_base = df_final.copy()
        perf_base["OA Flag"] = perf_base["Official Account"].apply(
            lambda x: "Official Account"
            if str(x).strip().lower() == "official account" else "Non Official Account"
        )

        perf_tbl = (
            perf_base
            .groupby(["Channel", "OA Flag"])
            .agg(Total_Engagement=("Engagement", "sum"),
                Number_of_Post=("Engagement", "size"))
            .reset_index()
        )
        perf_tbl["Avg. Engagement"] = (
            perf_tbl["Total_Engagement"] / perf_tbl["Number_of_Post"]
        ).round(2)

        perf_tbl = perf_tbl.rename(columns={
            "Channel": "Channel",
            "OA Flag": "Account Type",
            "Total_Engagement": "Total Engagement",
            "Number_of_Post": "Number of Post"
        })




        ### --- POSISI BLOK DI SHEET SUMMARY ---------------------------------
        summary_col_count = df_summary.shape[1]           # kolom blok Overview
        sov_start_col     = summary_col_count + 2         # +2 kolom spasi

        sov_col_letter    = get_column_letter(sov_start_col + 1)

        # Blok Top-Author ditempatkan setelah SOV + 2 kolom spasi
        top_buzz_start_col = sov_start_col + df_sov.shape[1] + 2
        top_post_start_col = top_buzz_start_col + len(top_buzz.columns) + 2

        buzz_col_letter = get_column_letter(top_buzz_start_col + 1)
        post_col_letter = get_column_letter(top_post_start_col + 1)



        # ----------------------------------------------------------------
        # === Hitung start-col untuk blok hashtag (2 kolom spasi) ---------
        # ----------------------------------------------------------------
        top_hash_buzz_start_col = top_post_start_col + len(top_post.columns) + 2
        top_hash_post_start_col = top_hash_buzz_start_col + len(top_hash_buzz.columns) + 2

        hash_buzz_col_letter = get_column_letter(top_hash_buzz_start_col + 1)
        hash_post_col_letter = get_column_letter(top_hash_post_start_col + 1)



        # ----------------------------------------------------------------
        # === Hitung start-col untuk blok Sentiment & Top-Post  =========   ### NEW
        # ----------------------------------------------------------------
        sent_start_col        = top_hash_post_start_col + len(top_hash_post.columns) + 2
        posbuzz_oa_start_col  = sent_start_col        + df_sent_pct.shape[1]         + 2
        posbuzz_nona_start_col= posbuzz_oa_start_col  + len(top_posbuzz_oa.columns)  + 2
        negbuzz_start_col     = posbuzz_nona_start_col+ len(top_posbuzz_nona.columns)+ 2

        sent_col_letter      = get_column_letter(sent_start_col + 1)

        posbuzz_oa_col_letter   = get_column_letter(posbuzz_oa_start_col   + 1)
        posbuzz_nona_col_letter = get_column_letter(posbuzz_nona_start_col + 1)
        negbuzz_col_letter      = get_column_letter(negbuzz_start_col      + 1)




        # ----------------------------------------------------------------
        # === Hitung start-col untuk top content dan top campaigns performance  =========   ### NEW
        # ----------------------------------------------------------------
        top_content_start_col   = negbuzz_start_col + len(top_negbuzz.columns) + 2
        perf_start_col          = top_content_start_col + len(top_content_buzz.columns) + 2

        content_col_letter = get_column_letter(top_content_start_col + 1)
        perf_col_letter    = get_column_letter(perf_start_col + 1)

        



        if not keep_raw_data and (df_final is None or df_final.empty):
            st.error("❌ Tidak ada data yang bisa disimpan. Hasil kosong dan tidak memilih simpan RAW data.")
        else:
            with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
                if keep_raw_data:
                    df_raw.to_excel(writer, sheet_name="RAW Data", index=False)
                if df_final is not None and not df_final.empty:
                    df_final.to_excel(writer, sheet_name="Process Data", index=False)
                if df_summary is not None and not df_summary.empty:
                    # Tambahkan sheet "Summary" di posisi kanan (terakhir)
                    df_summary.to_excel(writer, sheet_name="Summary", index=False, startrow=1, startcol=1)
                    # --- tulis Share Of Voice ---
                    df_sov.to_excel(
                        writer, sheet_name="Summary",
                        index=False,
                        startrow=1,           # baris 2 (row-index 1) → data
                        startcol=sov_start_col
                    )


                    # --- tulis tabel Top 10 Author -----------------------------------
                    top_buzz.to_excel(writer, sheet_name="Summary",
                        index=False, startrow=1, startcol=top_buzz_start_col)

                    top_post.to_excel(writer, sheet_name="Summary",
                        index=False, startrow=1, startcol=top_post_start_col)
                    
                    # --- tulis Top Hashtag -----------------------------------------
                    top_hash_buzz.to_excel(
                        writer, sheet_name="Summary",
                        index=False, startrow=1, startcol=top_hash_buzz_start_col
                    )
                    top_hash_post.to_excel(
                        writer, sheet_name="Summary",
                        index=False, startrow=1, startcol=top_hash_post_start_col
                    )


                    # --- tulis Sentiment Percentage --------------------------------      ### NEW
                    df_sent_pct.to_excel(
                        writer, sheet_name="Summary",
                        index=False, startrow=1, startcol=sent_start_col
                    )

                    # --- tulis Top 10 Positive & Negative Post by Buzz -------------      ### NEW
                    top_posbuzz_oa.to_excel(
                        writer, sheet_name="Summary",
                        index=False, startrow=1, startcol=posbuzz_oa_start_col
                    )
                    top_posbuzz_nona.to_excel(
                        writer, sheet_name="Summary",
                        index=False, startrow=1, startcol=posbuzz_nona_start_col
                    )
                    top_negbuzz.to_excel(
                        writer, sheet_name="Summary",
                        index=False, startrow=1, startcol=negbuzz_start_col
                    )


                    # --- tulis Top Content by Buzz ----------------------------------
                    top_content_buzz.to_excel(
                        writer, sheet_name="Summary",
                        index=False, startrow=1, startcol=top_content_start_col
                    )

                    # --- tulis Top Campaigns Performance ----------------------------
                    perf_tbl.to_excel(
                        writer, sheet_name="Summary",
                        index=False, startrow=1, startcol=perf_start_col
                    )



                    # =========================================================
                    # ===  AFTER WRITING  :  formatting via openpyxl
                    # =========================================================
                    wb = writer.book
                    ws = wb["Summary"]

                    bold = Font(bold=True)

                    # Judul blok
                    ws["B1"].value              = "Data Collection Overview"
                    ws["B1"].font               = bold

                    ws[f"{sov_col_letter}1"].value = "Share Of Voice"
                    ws[f"{sov_col_letter}1"].font  = bold

                    ws[f"{buzz_col_letter}1"].value = "Top 10 Author by Buzz"
                    ws[f"{buzz_col_letter}1"].font  = bold

                    ws[f"{post_col_letter}1"].value = "Top 10 Author by Post"
                    ws[f"{post_col_letter}1"].font  = bold

                    # Judul tabel hashtag
                    ws[f"{hash_buzz_col_letter}1"] = "Top Hashtag by Buzz"
                    ws[f"{hash_buzz_col_letter}1"].font = bold

                    ws[f"{hash_post_col_letter}1"] = "Top Hashtag by Post"
                    ws[f"{hash_post_col_letter}1"].font = bold


                    # Judul blok Sentiment & Top-Post                                  ### NEW
                    ws[f"{sent_col_letter}1"]   = "Sentiment Percentage"
                    ws[f"{sent_col_letter}1"].font = bold

                    ws[f"{posbuzz_oa_col_letter}1"]   = "Top 10 Post Positive by Buzz (Official Account)"
                    ws[f"{posbuzz_nona_col_letter}1"] = "Top 10 Post Positive by Buzz (Non Official Account)"
                    ws[f"{negbuzz_col_letter}1"]      = "Top 10 Post Negative by Buzz"

                    # Bold semua
                    for cell in [f"{posbuzz_oa_col_letter}1", f"{posbuzz_nona_col_letter}1", f"{negbuzz_col_letter}1"]:
                        ws[cell].font = bold


                    ws[f"{content_col_letter}1"] = "Top Content by Buzz"
                    ws[f"{content_col_letter}1"].font = bold

                    ws[f"{perf_col_letter}1"] = "Top Campaigns Performance"
                    ws[f"{perf_col_letter}1"].font = bold



                    # === AUTO-FIT width semua kolom di sheet Summary ====================
                    MAX_CHAR_WIDTH = 82.33  # Excel column width ≈ 500px

                    for col_cells in ws.columns:
                        max_len = 0
                        col_letter = col_cells[0].column_letter
                        for c in col_cells:
                            if c.value is not None:
                                max_len = max(max_len, len(str(c.value)))
                                # pastikan alignment tidak wrap
                                c.alignment = Alignment(wrap_text=False)
                        # Set column width dengan batas maksimum
                        ws.column_dimensions[col_letter].width = min(max_len + 2, MAX_CHAR_WIDTH)



        # Simpan nama file ke session state agar bisa diakses oleh tombol download
        st.session_state["download_filename"] = output_filename


        end_time = time.time()
        minutes, seconds = divmod(end_time - start_time, 60)

        # === Hitung durasi proses
        duration_seconds = end_time - start_time
        hours, remainder = divmod(duration_seconds, 3600)
        minutes, seconds = divmod(remainder, 60)


        # Simpan ke session_state agar tidak hilang saat rerun
        #st.session_state["output_excel_bytes"] = output_buffer.getvalue()
        st.session_state["download_filename"] = output_filename
        st.session_state["summary_df"] = summary_combined
        st.session_state["df_final"] = df_final


        # Simpan Noise Tag Summary
        if "Noise Tag" in df_processed.columns:
            noise_summary = df_processed["Noise Tag"].astype(str).str.replace(r"\.0$", "", regex=True).value_counts().reset_index()
            noise_summary.columns = ["Noise Tag", "Jumlah"]
            description_map = {
                "0": "Valid Content",
                "1": "Official Brand",
                "2": "Exclude Content (noise)",
                "3": "Need QC"
            }
            noise_summary["Description"] = noise_summary["Noise Tag"].astype(str).map(description_map)
            st.session_state["noise_summary_df"] = noise_summary

        # Simpan durasi proses
        st.session_state["process_duration"] = f"{int(hours)} jam {int(minutes)} menit {int(seconds)} detik"

        #Finish
        update_progress(14, 14, "✅ Process Completed.")



        with open(st.session_state["download_filename"], "rb") as file:
            st.download_button(
                label="⬇️ Download Hasil Excel",
                data=file.read(),
                file_name=st.session_state["download_filename"],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )


    # Tampilkan durasi proses
    if "process_duration" in st.session_state:
        st.info(f"🕒 Proses ini berjalan selama {st.session_state['process_duration']}.")

    if "Output Value" in summary_df.columns:
        summary_df["Output Value"] = summary_df["Output Value"].astype(str).str.replace(r"\.0$", "", regex=True)
    if summary_df.empty:
        st.warning("⚠️ Tidak ada rule yang diterapkan. Summary kosong.")
    
    # Tampilkan Summary Execution Report jika ada
    if "summary_df" in st.session_state:
        st.subheader("📊 Summary Execution Report")
        with st.expander("**🧮 Lihat Summary Execution Report**"):
            summary_df = st.session_state["summary_df"]
            summary_df["Output Value"] = summary_df["Output Value"].astype(str).str.replace(r"\.0$", "", regex=True)
            summary_sorted = summary_df.sort_values(by="Priority", ascending=False)
            st.dataframe(summary_sorted[[
                "Priority", "Matching Column", "Matching Value", "Matching Type",
                "Channel", "Affected Rows", "Output Column", "Output Value"
            ]])

            # Tambahkan Ringkasan Noise Tag jika ada
            if "noise_summary_df" in st.session_state:
                st.markdown("**🧮 Ringkasan Noise Tag:**")
                st.dataframe(st.session_state["noise_summary_df"]) 
            
            if "df_final" in st.session_state:
                if "Rules Affected" in st.session_state["df_final"].columns:
                    st.subheader("📌 Rules Affected Preview")
                    st.dataframe(
                        st.session_state["df_final"][["Rules Affected", "Rules Affected Words"]]
                    )


            if "df_final" in st.session_state and "Creator Type" in st.session_state["df_final"].columns:
                st.markdown("**🧬 Ringkasan Creator Type:**")
                creator_summary = (
                    st.session_state["df_final"]["Creator Type"]
                    .fillna("")
                    .replace("", "(blank)")
                    .value_counts()
                    .reset_index()
                )
                creator_summary.columns = ["Creator Type", "Jumlah"]
                # Urutkan: selain (blank) dulu, lalu (blank) di bawah
                creator_summary["is_blank"] = creator_summary["Creator Type"] == "(blank)"
                creator_summary = creator_summary.sort_values(by=["is_blank", "Jumlah"], ascending=[True, False]).drop(columns="is_blank")
                st.dataframe(creator_summary)


    #st.success(f"📄 Log file saved as: `logs/{log_filename}`")


else:
    st.stop()